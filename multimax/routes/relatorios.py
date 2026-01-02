from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session
from flask_login import login_required, current_user
from .. import db
from ..models import (
    Produto, Historico, Collaborator,
    User, JobRole, Holiday, Recipe, IngredientCatalog, CleaningTask, CleaningChecklistTemplate,
    TimeOffRecord
)
from datetime import datetime, date
from zoneinfo import ZoneInfo
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Mapeamento de tabelas para exportação/importação genérica
TABLE_MAP = {
    'users': User,
    'produtos': Produto,
    'colaboradores': Collaborator,
    'cargos': JobRole,
    'feriados': Holiday,
    'receitas': Recipe,
    'ingredientes_catalogo': IngredientCatalog,
    'tarefas_limpeza': CleaningTask,
    'checklist_templates': CleaningChecklistTemplate
}

def _convert_value(value, col_type):
    """Converte valores do Excel para tipos do banco de dados"""
    if value is None:
        return None
    s_type = str(col_type).upper()
    if 'DATE' in s_type or 'TIME' in s_type:
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value)
            except (ValueError, TypeError):
                try:
                    return datetime.strptime(value, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
    if 'BOOLEAN' in s_type:
        if isinstance(value, str):
            return value.lower() in ('true', '1', 't', 'y', 'yes')
        return bool(value)
    return value

bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')

@bp.route('/')
@login_required
def index():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para acessar esta página.', 'danger')
        return redirect(url_for('home.index'))
    return render_template('relatorios.html', active_page='relatorios')

@bp.route('/exportar/produtos')
@login_required
def exportar_produtos():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    produtos = Produto.query.order_by(Produto.nome.asc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Código', 'Nome', 'Quantidade', 'Estoque Mínimo', 'Preço Custo', 'Preço Venda', 'Data Validade', 'Lote']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, produto in enumerate(produtos, 2):
        data = [
            produto.codigo,
            produto.nome,
            produto.quantidade,
            produto.estoque_minimo,
            produto.preco_custo or 0,
            produto.preco_venda or 0,
            produto.data_validade.strftime('%d/%m/%Y') if produto.data_validade else '',
            produto.lote or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"produtos_multimax_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/exportar/historico')
@login_required
def exportar_historico():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    historico = Historico.query.order_by(Historico.data.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Data/Hora', 'Produto', 'Ação', 'Quantidade', 'Detalhes', 'Usuário']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, h in enumerate(historico, 2):
        data = [
            h.data.strftime('%d/%m/%Y %H:%M:%S') if h.data else '',
            h.product_name or '',
            'Entrada' if h.action == 'entrada' else 'Saída',
            h.quantidade or 0,
            h.details or '',
            h.usuario or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"historico_multimax_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/importar/produtos', methods=['GET', 'POST'])
@login_required
def importar_produtos():
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar produtos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_produtos'))
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_produtos'))
        
        if not arquivo.filename.endswith('.xlsx'):
            flash('Formato inválido. Use arquivos .xlsx', 'danger')
            return redirect(url_for('relatorios.importar_produtos'))
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            importados = 0
            atualizados = 0
            erros = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                codigo = str(row[0]).strip()
                nome = str(row[1]).strip() if row[1] else ''
                
                if not nome:
                    erros += 1
                    continue
                
                try:
                    quantidade = int(row[2]) if row[2] else 0
                except Exception:
                    quantidade = 0
                
                try:
                    estoque_minimo = int(row[3]) if row[3] else 0
                except Exception:
                    estoque_minimo = 0
                
                try:
                    preco_custo = float(row[4]) if row[4] else 0.0
                except Exception:
                    preco_custo = 0.0
                
                try:
                    preco_venda = float(row[5]) if row[5] else 0.0
                except Exception:
                    preco_venda = 0.0
                
                produto = Produto.query.filter_by(codigo=codigo).first()
                
                if produto:
                    produto.nome = nome
                    produto.quantidade = quantidade
                    produto.estoque_minimo = estoque_minimo
                    produto.preco_custo = preco_custo
                    produto.preco_venda = preco_venda
                    atualizados += 1
                else:
                    novo_produto = Produto()
                    novo_produto.codigo = codigo
                    novo_produto.nome = nome
                    novo_produto.quantidade = quantidade
                    novo_produto.estoque_minimo = estoque_minimo
                    novo_produto.preco_custo = preco_custo
                    novo_produto.preco_venda = preco_venda
                    db.session.add(novo_produto)
                    importados += 1
            
            db.session.commit()
            flash(f'Importação concluída! {importados} novos produtos, {atualizados} atualizados, {erros} erros.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao importar: {e}', 'danger')
        
        return redirect(url_for('relatorios.index'))
    
    return render_template('importar_produtos.html', active_page='relatorios')

@bp.route('/modelo/produtos')
@login_required
def modelo_produtos():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para baixar modelos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Produtos"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Código', 'Nome', 'Quantidade', 'Estoque Mínimo', 'Preço Custo', 'Preço Venda']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    exemplo = ['CX-0001', 'Produto Exemplo', 100, 10, 15.50, 25.00]
    for col, value in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='modelo_produtos.xlsx', as_attachment=True)

@bp.route('/importar/horas-extras', methods=['GET', 'POST'])
@login_required
def importar_horas_extras():
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar horas extras.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_horas_extras'))
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_horas_extras'))
        
        if not arquivo.filename.endswith('.xlsx'):
            flash('Formato inválido. Use arquivos .xlsx', 'danger')
            return redirect(url_for('relatorios.importar_horas_extras'))
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            importados = 0
            erros = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                # Colunas: ID Colaborador, Data (DD/MM/YYYY), Horas, Motivo
                try:
                    collaborator_id = int(row[0]) if row[0] else None
                    if not collaborator_id:
                        erros += 1
                        continue
                    
                    # Verificar se colaborador existe
                    collaborator = Collaborator.query.get(collaborator_id)
                    if not collaborator:
                        erros += 1
                        continue
                    
                    # Processar data
                    data_str = str(row[1]).strip() if row[1] else ''
                    if not data_str:
                        erros += 1
                        continue
                    
                    try:
                        # Tentar diferentes formatos de data
                        if '/' in data_str:
                            parts = data_str.split('/')
                            if len(parts) == 3:
                                data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            else:
                                erros += 1
                                continue
                        elif '-' in data_str:
                            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                        else:
                            erros += 1
                            continue
                    except (ValueError, TypeError, IndexError):
                        erros += 1
                        continue
                    
                    # Processar horas
                    try:
                        horas = float(row[2]) if row[2] else 0.0
                        if horas == 0:
                            erros += 1
                            continue
                    except (ValueError, TypeError):
                        erros += 1
                        continue
                    
                    # Motivo (opcional)
                    motivo = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    
                    # Criar registro
                    entry = TimeOffRecord()
                    entry.collaborator_id = collaborator_id
                    entry.date = data_obj
                    entry.record_type = 'horas'
                    entry.hours = horas
                    entry.notes = motivo or 'Importado via Excel'
                    entry.origin = 'excel'
                    entry.created_by = current_user.username if current_user.is_authenticated else 'sistema'
                    db.session.add(entry)
                    importados += 1
                    
                except Exception as e:
                    erros += 1
                    continue
            
            db.session.commit()
            flash(f'Importação concluída! {importados} registros de horas extras importados, {erros} erros.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao importar: {e}', 'danger')
        
        return redirect(url_for('relatorios.index'))
    
    return render_template('importar_horas_extras.html', active_page='relatorios')

@bp.route('/modelo/horas-extras')
@login_required
def modelo_horas_extras():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para baixar modelos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Horas Extras"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Horas', 'Motivo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    exemplo = [1, '01/01/2025', 2.5, 'Trabalho extra']
    for col, value in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='modelo_horas_extras.xlsx', as_attachment=True)

@bp.route('/importar/folgas-adicionais', methods=['GET', 'POST'])
@login_required
def importar_folgas_adicionais():
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar folgas adicionais.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_folgas_adicionais'))
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_folgas_adicionais'))
        
        if not arquivo.filename.endswith('.xlsx'):
            flash('Formato inválido. Use arquivos .xlsx', 'danger')
            return redirect(url_for('relatorios.importar_folgas_adicionais'))
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            importados = 0
            erros = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                # Colunas: ID Colaborador, Data (DD/MM/YYYY), Dias, Origem, Observações
                try:
                    collaborator_id = int(row[0]) if row[0] else None
                    if not collaborator_id:
                        erros += 1
                        continue
                    
                    # Verificar se colaborador existe
                    collaborator = Collaborator.query.get(collaborator_id)
                    if not collaborator:
                        erros += 1
                        continue
                    
                    # Processar data
                    data_str = str(row[1]).strip() if row[1] else ''
                    if not data_str:
                        erros += 1
                        continue
                    
                    try:
                        # Tentar diferentes formatos de data
                        if '/' in data_str:
                            parts = data_str.split('/')
                            if len(parts) == 3:
                                data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            else:
                                erros += 1
                                continue
                        elif '-' in data_str:
                            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                        else:
                            erros += 1
                            continue
                    except (ValueError, TypeError, IndexError):
                        erros += 1
                        continue
                    
                    # Processar dias
                    try:
                        dias = int(row[2]) if row[2] else 0
                        if dias <= 0:
                            erros += 1
                            continue
                    except (ValueError, TypeError):
                        erros += 1
                        continue
                    
                    # Origem (opcional)
                    origem = str(row[3]).strip() if len(row) > 3 and row[3] else 'excel'
                    
                    # Observações (opcional)
                    observacoes = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    
                    # Criar registro
                    credit = TimeOffRecord()
                    credit.collaborator_id = collaborator_id
                    credit.date = data_obj
                    credit.record_type = 'folga_adicional'
                    credit.days = dias
                    credit.origin = origem
                    credit.notes = observacoes
                    credit.created_by = current_user.username if current_user.is_authenticated else 'sistema'
                    db.session.add(credit)
                    importados += 1
                    
                except Exception as e:
                    erros += 1
                    continue
            
            db.session.commit()
            flash(f'Importação concluída! {importados} registros de folgas adicionais importados, {erros} erros.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao importar: {e}', 'danger')
        
        return redirect(url_for('relatorios.index'))
    
    return render_template('importar_folgas_adicionais.html', active_page='relatorios')

@bp.route('/modelo/folgas-adicionais')
@login_required
def modelo_folgas_adicionais():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para baixar modelos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Folgas Adicionais"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Dias', 'Origem', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    exemplo = [1, '01/01/2025', 1, 'excel', 'Folga adicional importada']
    for col, value in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='modelo_folgas_adicionais.xlsx', as_attachment=True)

@bp.route('/importar/folgas-usadas', methods=['GET', 'POST'])
@login_required
def importar_folgas_usadas():
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar folgas usadas.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_folgas_usadas'))
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_folgas_usadas'))
        
        if not arquivo.filename.endswith('.xlsx'):
            flash('Formato inválido. Use arquivos .xlsx', 'danger')
            return redirect(url_for('relatorios.importar_folgas_usadas'))
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            importados = 0
            erros = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                # Colunas: ID Colaborador, Data (DD/MM/YYYY), Dias Usados, Observações
                try:
                    collaborator_id = int(row[0]) if row[0] else None
                    if not collaborator_id:
                        erros += 1
                        continue
                    
                    # Verificar se colaborador existe
                    collaborator = Collaborator.query.get(collaborator_id)
                    if not collaborator:
                        erros += 1
                        continue
                    
                    # Processar data
                    data_str = str(row[1]).strip() if row[1] else ''
                    if not data_str:
                        erros += 1
                        continue
                    
                    try:
                        # Tentar diferentes formatos de data
                        if '/' in data_str:
                            parts = data_str.split('/')
                            if len(parts) == 3:
                                data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            else:
                                erros += 1
                                continue
                        elif '-' in data_str:
                            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                        else:
                            erros += 1
                            continue
                    except (ValueError, TypeError, IndexError):
                        erros += 1
                        continue
                    
                    # Processar dias usados
                    try:
                        dias_usados = int(row[2]) if row[2] else 0
                        if dias_usados <= 0:
                            erros += 1
                            continue
                    except (ValueError, TypeError):
                        erros += 1
                        continue
                    
                    # Observações (opcional)
                    observacoes = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    
                    # Criar registro
                    assignment = TimeOffRecord()
                    assignment.collaborator_id = collaborator_id
                    assignment.date = data_obj
                    assignment.record_type = 'folga_usada'
                    assignment.days = dias_usados
                    assignment.notes = observacoes
                    assignment.origin = 'excel'
                    assignment.created_by = current_user.username if current_user.is_authenticated else 'sistema'
                    db.session.add(assignment)
                    importados += 1
                    
                except Exception as e:
                    erros += 1
                    continue
            
            db.session.commit()
            flash(f'Importação concluída! {importados} registros de folgas usadas importados, {erros} erros.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao importar: {e}', 'danger')
        
        return redirect(url_for('relatorios.index'))
    
    return render_template('importar_folgas_usadas.html', active_page='relatorios')

@bp.route('/modelo/folgas-usadas')
@login_required
def modelo_folgas_usadas():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para baixar modelos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Folgas Usadas"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Dias Usados', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    exemplo = [1, '01/01/2025', 1, 'Folga usada importada']
    for col, value in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='modelo_folgas_usadas.xlsx', as_attachment=True)

@bp.route('/exportar/horas-extras')
@login_required
def exportar_horas_extras():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    horas_extras = TimeOffRecord.query.filter(TimeOffRecord.record_type == 'horas').order_by(TimeOffRecord.date.desc(), TimeOffRecord.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Horas', 'Motivo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, entry in enumerate(horas_extras, 2):
        data = [
            entry.collaborator_id,
            entry.date.strftime('%d/%m/%Y') if entry.date else '',
            float(entry.hours or 0.0),
            entry.notes or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"horas_extras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/exportar/folgas-adicionais')
@login_required
def exportar_folgas_adicionais():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    folgas_adicionais = TimeOffRecord.query.filter(TimeOffRecord.record_type == 'folga_adicional').order_by(TimeOffRecord.date.desc(), TimeOffRecord.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Folgas Adicionais"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Dias', 'Origem', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, credit in enumerate(folgas_adicionais, 2):
        data = [
            credit.collaborator_id,
            credit.date.strftime('%d/%m/%Y') if credit.date else '',
            int(credit.days or 0),
            credit.origin or '',
            credit.notes or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"folgas_adicionais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/exportar/folgas-usadas')
@login_required
def exportar_folgas_usadas():
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    folgas_usadas = TimeOffRecord.query.filter(TimeOffRecord.record_type == 'folga_usada').order_by(TimeOffRecord.date.desc(), TimeOffRecord.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Folgas Usadas"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Dias Usados', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, assignment in enumerate(folgas_usadas, 2):
        data = [
            assignment.collaborator_id,
            assignment.date.strftime('%d/%m/%Y') if assignment.date else '',
            int(assignment.days or 0),
            assignment.notes or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"folgas_usadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/exportar/registros-unificados')
@login_required
def exportar_registros_unificados():
    """Exporta registros unificados de horas e folgas"""
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para exportar dados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    collaborator_id = request.args.get('collaborator_id', type=int)
    data_inicio = request.args.get('inicio', '').strip()
    data_fim = request.args.get('fim', '').strip()
    
    di = None
    df = None
    if data_inicio:
        try:
            di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        except Exception:
            pass
    if data_fim:
        try:
            df = datetime.strptime(data_fim, '%Y-%m-%d').date()
        except Exception:
            pass
    
    q = TimeOffRecord.query
    if collaborator_id:
        q = q.filter(TimeOffRecord.collaborator_id == collaborator_id)
    if di:
        q = q.filter(TimeOffRecord.date >= di)
    if df:
        q = q.filter(TimeOffRecord.date <= df)
    
    records = q.order_by(TimeOffRecord.date.desc(), TimeOffRecord.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros Unificados"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Tipo', 'Horas', 'Dias', 'Valor Pago (R$)', 'Taxa por Dia (R$)', 'Origem', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, record in enumerate(records, 2):
        data = [
            record.collaborator_id,
            record.date.strftime('%d/%m/%Y') if record.date else '',
            record.record_type or '',
            float(record.hours) if record.hours is not None else '',
            int(record.days) if record.days is not None else '',
            float(record.amount_paid) if record.amount_paid is not None else '',
            float(record.rate_per_day) if record.rate_per_day is not None else '',
            record.origin or '',
            record.notes or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"registros_unificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)

@bp.route('/importar/registros-unificados', methods=['GET', 'POST'])
@login_required
def importar_registros_unificados():
    """Importa registros unificados de horas e folgas com tratamento detalhado de erros"""
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar registros unificados.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_registros_unificados'))
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('relatorios.importar_registros_unificados'))
        
        if not arquivo.filename.endswith('.xlsx'):
            flash('Formato inválido. Use arquivos .xlsx', 'danger')
            return redirect(url_for('relatorios.importar_registros_unificados'))
        
        importados = 0
        erros = []
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                
                try:
                    # Colunas: ID Colaborador, Data (DD/MM/YYYY), Tipo, Horas, Dias, Valor Pago, Taxa por Dia, Origem, Observações
                    collaborator_id = None
                    if row[0]:
                        try:
                            collaborator_id = int(row[0])
                        except (ValueError, TypeError):
                            erros.append(f'Linha {row_num}: ID do colaborador inválido: "{row[0]}"')
                            continue
                    
                    if not collaborator_id:
                        erros.append(f'Linha {row_num}: ID do colaborador é obrigatório')
                        continue
                    
                    # Verificar se colaborador existe
                    collaborator = Collaborator.query.get(collaborator_id)
                    if not collaborator:
                        erros.append(f'Linha {row_num}: Colaborador com ID {collaborator_id} não encontrado')
                        continue
                    
                    # Processar data
                    data_str = str(row[1]).strip() if row[1] else ''
                    if not data_str:
                        erros.append(f'Linha {row_num}: Data é obrigatória')
                        continue
                    
                    try:
                        if '/' in data_str:
                            parts = data_str.split('/')
                            if len(parts) == 3:
                                data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            else:
                                erros.append(f'Linha {row_num}: Formato de data inválido: "{data_str}". Use DD/MM/YYYY')
                                continue
                        elif '-' in data_str:
                            data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
                        else:
                            erros.append(f'Linha {row_num}: Formato de data inválido: "{data_str}"')
                            continue
                    except (ValueError, TypeError, IndexError) as e:
                        erros.append(f'Linha {row_num}: Erro ao processar data "{data_str}": {str(e)}')
                        continue
                    
                    # Processar tipo
                    record_type = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ''
                    if not record_type:
                        erros.append(f'Linha {row_num}: Tipo de registro é obrigatório')
                        continue
                    
                    if record_type not in ('horas', 'folga_adicional', 'folga_usada', 'conversao'):
                        erros.append(f'Linha {row_num}: Tipo inválido: "{record_type}". Use: horas, folga_adicional, folga_usada ou conversao')
                        continue
                    
                    # Criar registro
                    record = TimeOffRecord()
                    record.collaborator_id = collaborator_id
                    record.date = data_obj
                    record.record_type = record_type
                    record.created_by = current_user.username
                    
                    # Processar campos específicos por tipo
                    if record_type == 'horas':
                        try:
                            hours = float(row[3]) if len(row) > 3 and row[3] is not None else None
                            if hours is None or hours == 0:
                                erros.append(f'Linha {row_num}: Horas devem ser informadas e diferentes de zero')
                                continue
                            record.hours = hours
                            record.notes = str(row[8]).strip() if len(row) > 8 and row[8] else 'Horas extras importadas'
                            record.origin = str(row[7]).strip() if len(row) > 7 and row[7] else 'excel'
                            
                            # Registro criado apenas na tabela unificada
                            
                        except (ValueError, TypeError) as e:
                            erros.append(f'Linha {row_num}: Erro ao processar horas: {str(e)}')
                            continue
                    
                    elif record_type == 'folga_adicional':
                        try:
                            days = int(row[4]) if len(row) > 4 and row[4] is not None else None
                            if days is None or days <= 0:
                                erros.append(f'Linha {row_num}: Dias devem ser informados e maiores que zero')
                                continue
                            record.days = days
                            record.notes = str(row[8]).strip() if len(row) > 8 and row[8] else 'Folga adicional importada'
                            record.origin = str(row[7]).strip() if len(row) > 7 and row[7] else 'excel'
                            
                            # Registro criado apenas na tabela unificada
                            
                        except (ValueError, TypeError) as e:
                            erros.append(f'Linha {row_num}: Erro ao processar dias: {str(e)}')
                            continue
                    
                    elif record_type == 'folga_usada':
                        try:
                            days = int(row[4]) if len(row) > 4 and row[4] is not None else None
                            if days is None or days <= 0:
                                erros.append(f'Linha {row_num}: Dias usados devem ser informados e maiores que zero')
                                continue
                            record.days = days
                            record.notes = str(row[8]).strip() if len(row) > 8 and row[8] else 'Folga usada importada'
                            record.origin = str(row[7]).strip() if len(row) > 7 and row[7] else 'excel'
                            
                            # Registro criado apenas na tabela unificada
                            
                        except (ValueError, TypeError) as e:
                            erros.append(f'Linha {row_num}: Erro ao processar dias usados: {str(e)}')
                            continue
                    
                    elif record_type == 'conversao':
                        try:
                            days = int(row[4]) if len(row) > 4 and row[4] is not None else None
                            amount_paid = float(row[5]) if len(row) > 5 and row[5] is not None else None
                            rate = float(row[6]) if len(row) > 6 and row[6] is not None else 65.0
                            
                            if days is None or days <= 0:
                                erros.append(f'Linha {row_num}: Dias convertidos devem ser informados e maiores que zero')
                                continue
                            if amount_paid is None:
                                erros.append(f'Linha {row_num}: Valor pago é obrigatório para conversões')
                                continue
                            
                            record.days = days
                            record.amount_paid = amount_paid
                            record.rate_per_day = rate
                            record.notes = str(row[8]).strip() if len(row) > 8 and row[8] else 'Conversão importada'
                            record.origin = str(row[7]).strip() if len(row) > 7 and row[7] else 'excel'
                            
                            # Registro criado apenas na tabela unificada
                            
                        except (ValueError, TypeError) as e:
                            erros.append(f'Linha {row_num}: Erro ao processar conversão: {str(e)}')
                            continue
                    
                    db.session.add(record)
                    importados += 1
                    
                except Exception as e:
                    erros.append(f'Linha {row_num}: Erro inesperado: {str(e)}')
                    continue
            
            db.session.commit()
            
            # Armazenar resultado na sessão para exibição detalhada
            session['import_result'] = {
                'importados': importados,
                'erros': erros,
                'total_linhas': len(erros) + importados
            }
            
            if erros:
                if importados > 0:
                    flash(f'Importação parcial: {importados} registro(s) importado(s) com sucesso, {len(erros)} erro(s) encontrado(s). Veja os detalhes abaixo.', 'warning')
                else:
                    flash(f'Importação falhou: {len(erros)} erro(s) encontrado(s). Nenhum registro foi importado. Veja os detalhes abaixo.', 'danger')
            else:
                flash(f'Importação concluída com sucesso! {importados} registro(s) importado(s).', 'success')
            
        except Exception as e:
            db.session.rollback()
            session['import_result'] = {
                'importados': 0,
                'erros': [f'Erro ao processar arquivo: {str(e)}'],
                'total_linhas': 0
            }
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
        
        return redirect(url_for('relatorios.resultado_importacao'))
    
    return render_template('importar_registros_unificados.html', active_page='relatorios')

@bp.route('/resultado-importacao')
@login_required
def resultado_importacao():
    """Exibe resultado detalhado da importação de registros unificados"""
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    resultado = session.pop('import_result', None)
    if not resultado:
        flash('Nenhum resultado de importação encontrado.', 'info')
        return redirect(url_for('relatorios.importar_registros_unificados'))
    
    return render_template('resultado_importacao.html', 
                         active_page='relatorios',
                         resultado=resultado)

@bp.route('/modelo/registros-unificados')
@login_required
def modelo_registros_unificados():
    """Gera modelo Excel para importação de registros unificados"""
    if current_user.nivel not in ['operador', 'admin', 'DEV']:
        flash('Você não tem permissão para baixar modelos.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Registros Unificados"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID Colaborador', 'Data (DD/MM/YYYY)', 'Tipo', 'Horas', 'Dias', 'Valor Pago (R$)', 'Taxa por Dia (R$)', 'Origem', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Exemplos
    exemplos = [
        [1, '01/01/2025', 'horas', 2.5, '', '', '', 'excel', 'Trabalho extra'],
        [1, '02/01/2025', 'folga_adicional', '', 1, '', '', 'manual', 'Folga adicional'],
        [1, '03/01/2025', 'folga_usada', '', 1, '', '', 'excel', 'Folga usada'],
        [1, '04/01/2025', 'conversao', '', 2, 130.00, 65.00, 'excel', 'Conversão em dinheiro']
    ]
    
    for row_num, exemplo in enumerate(exemplos, 2):
        for col, value in enumerate(exemplo, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='modelo_registros_unificados.xlsx', as_attachment=True)

@bp.route('/exportar/tabela/<table_name>', methods=['GET'])
@login_required
def exportar_tabela(table_name):
    """Exporta qualquer tabela do sistema para Excel"""
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem exportar tabelas.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    model = TABLE_MAP.get(table_name)
    if not model:
        flash('Tabela não encontrada.', 'danger')
        return redirect(url_for('relatorios.index'))
    
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = table_name
        
        # Headers
        columns = [c.key for c in model.__table__.columns]
        ws.append(columns)
        
        # Data
        records = model.query.all()
        for record in records:
            row = []
            for col in columns:
                val = getattr(record, col)
                if isinstance(val, (datetime, date)):
                    val = val.isoformat()
                row.append(val)
            ws.append(row)
            
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'Erro ao exportar: {e}', 'danger')
        return redirect(url_for('relatorios.index'))

@bp.route('/importar/tabela/<table_name>', methods=['POST'])
@login_required
def importar_tabela(table_name):
    """Importa dados de Excel para qualquer tabela do sistema"""
    if current_user.nivel not in ('admin', 'DEV'):
        flash('Apenas administradores podem importar tabelas.', 'danger')
        return redirect(url_for('relatorios.index'))
        
    model = TABLE_MAP.get(table_name)
    if not model:
        flash('Tabela não encontrada.', 'danger')
        return redirect(url_for('relatorios.index'))
        
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('relatorios.index'))
        
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado.', 'danger')
        return redirect(url_for('relatorios.index'))
        
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        rows = list(ws.rows)
        if not rows:
            flash('Arquivo vazio.', 'warning')
            return redirect(url_for('relatorios.index'))
            
        headers = [c.value for c in rows[0]]
        column_map = {c.key: c.type for c in model.__table__.columns}
        
        added = 0
        updated = 0
        
        for row in rows[1:]:
            values = [c.value for c in row]
            row_data = dict(zip(headers, values))
            
            # Remove keys not in model
            row_data = {k: v for k, v in row_data.items() if k in column_map}
            
            # Check for ID
            record_id = row_data.get('id')
            record = None
            if record_id:
                record = model.query.get(record_id)
            
            if record:
                # Update
                for k, v in row_data.items():
                    if k == 'id': 
                        continue
                    val = _convert_value(v, column_map.get(k))
                    setattr(record, k, val)
                updated += 1
            else:
                # Create
                if 'id' in row_data and not row_data['id']:
                    del row_data['id']
                
                # Convert values
                for k in row_data:
                    row_data[k] = _convert_value(row_data[k], column_map.get(k))
                    
                record = model(**row_data)
                db.session.add(record)
                added += 1
                
        db.session.commit()
        flash(f'Importação concluída: {added} adicionados, {updated} atualizados.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro na importação: {str(e)}', 'danger')
        
    return redirect(url_for('relatorios.index'))
